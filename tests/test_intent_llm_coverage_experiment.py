import csv
import io
import json
import tempfile
import unittest
from pathlib import Path

from agent.memory_guided_workflow.experiments.intent_llm_coverage_experiment import (
    BATCH_COLUMNS,
    COMPARISON_TABLE_COLUMNS,
    IntentLLMCoverageExperiment,
    coverage_result_to_batch_rows,
    coverage_result_to_comparison_row,
    load_existing_coverage_tables,
    load_gold_results,
    load_json_records,
    load_intent_tool_ids,
    load_intent_tools,
    load_intents,
    load_user_requests,
    main,
    load_qwen_results,
    render_markdown_table,
    run_batch,
    run_comparison_table_batch,
    write_batch_table,
    write_comparison_table,
)


class FakeLLMClient:
    def __init__(self, payload):
        self.payload = payload
        self.messages = []

    def chat(self, messages):
        self.messages = messages
        return json.dumps(self.payload, ensure_ascii=False)


class RawFakeLLMClient:
    def __init__(self, raw_text):
        self.raw_text = raw_text
        self.messages = []

    def chat(self, messages):
        self.messages = messages
        return self.raw_text


class RaisingLLMClient:
    def __init__(self, exc):
        self.exc = exc

    def chat(self, messages):
        raise self.exc


class PromptAwareFakeLLMClient:
    def chat(self, messages):
        prompt = messages[-1]["content"]
        if "Second request" in prompt:
            intent = "AnalysisSentimentOfText"
            phrase = "sentiment"
        else:
            intent = "SummarizeTextToShorterVersion"
            phrase = "summary"
        return json.dumps(
            {
                "covered_intents": [
                    {
                        "intent": intent,
                        "coverage_type": "direct",
                        "confidence": 0.95,
                        "matched_request_phrase": phrase,
                        "matched_intent_term": phrase,
                        "reason": "The request directly asks for this operation.",
                    }
                ]
            },
            ensure_ascii=False,
        )


class CountingPromptAwareFakeLLMClient(PromptAwareFakeLLMClient):
    def __init__(self):
        self.prompts = []

    def chat(self, messages):
        self.prompts.append(messages[-1]["content"])
        return super().chat(messages)


MULTIMEDIA_COVERAGE_PROMPT_RULES = [
    "URL phrases decompose by role: direct resource, embedded link in text, or link extraction target. This is coverage, not execution-order inference.",
    "For rewrite-like intents, distinguish the requested object and modifier: article-level objects and modifiers like unique version, synonyms, syntax changes, or plagiarism avoidance are stronger evidence for article-spinning intents; text-level objects and modifiers like different words, same meaning, paraphrase, or rephrase are stronger evidence for text paraphrasing intents.",
    "If an explicit http(s) URL is the source resource to process, cover the matching URL download intent from url_download_candidates as strongly_implied.",
    "If the URL is embedded inside a larger text/message and must be used later, cover URL extraction plus the matching URL download intent. Cover URL extraction alone only when the request asks to extract, find, list, or return URLs/links themselves.",
    "Image visualization: image/images/picture/illustration/visualize can cover image-related intents. Existing images imply image search; creating or visualizing an idea as images implies image generation.",
    "If the request provides non-English text, explicitly says the input text is written in another language, or asks for the result in a different language, cover the matching translation intent. Use strongly_implied when translation is needed for understanding before another content-level operation.",
]


class TestIntentLLMCoverageExperiment(unittest.TestCase):
    def test_covers_intents_from_synonym_matches(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=[
                "SummarizeTextToShorterVersion",
                "ParaphraseTextUsingDifferentWords",
                "AnalysisSentimentOfText",
            ],
            intent_tool_ids={
                "SummarizeTextToShorterVersion": ["Text Summarizer"],
                "ParaphraseTextUsingDifferentWords": ["Text Paraphraser"],
                "AnalysisSentimentOfText": ["Text Sentiment Analysis"],
            },
            llm_client=FakeLLMClient(
                {
                    "covered_intents": [
                        {
                            "intent": "SummarizeTextToShorterVersion",
                            "coverage_type": "synonym",
                            "confidence": 0.96,
                            "matched_request_phrase": "summarized version",
                            "matched_intent_term": "summarize",
                            "reason": "The phrase 'summarized version' is a morphological synonym of summarize.",
                        },
                        {
                            "intent": "AnalysisSentimentOfText",
                            "coverage_type": "direct",
                            "confidence": 0.98,
                            "matched_request_phrase": "sentiment",
                            "matched_intent_term": "sentiment",
                            "reason": "The request explicitly asks to know the sentiment.",
                        },
                    ]
                }
            ),
        )

        result = experiment.run("Find a summarized version and know the sentiment.")

        self.assertEqual(
            [row.intent for row in result.covered_intents],
            ["SummarizeTextToShorterVersion", "AnalysisSentimentOfText"],
        )
        markdown = render_markdown_table(result)
        self.assertIn("SummarizeTextToShorterVersion", markdown)
        self.assertIn("Text Summarizer", markdown)
        self.assertIn("Text Sentiment Analysis", markdown)
        self.assertNotIn("summarized version", markdown)
        self.assertNotIn("tool_id", markdown)
        self.assertNotIn("desc", markdown)
        self.assertEqual(result.covered_intents[0].tool_ids, ["Text Summarizer"])

    def test_skips_unknown_intent(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["SummarizeTextToShorterVersion"],
            llm_client=FakeLLMClient({"covered_intents": [{"intent": "UnknownIntent"}]}),
        )

        result = experiment.run("Summarize this text.")

        self.assertEqual(result.covered_intents, [])
        self.assertEqual(len(result.warnings), 1)

    def test_rejects_intents_outside_allowed_list(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["TranscribeAudioToText", "AnswerQuestionByRetrieveOrSearch"],
            intent_tool_ids={
                "TranscribeAudioToText": ["Automatic Speech Recognition"],
                "AnswerQuestionByRetrieveOrSearch": ["Question Answering"],
            },
            llm_client=FakeLLMClient({"covered_intents": [{"intent": "AnswerQuestionOnAudio"}]}),
        )

        result = experiment.run("Answer the question from this audio.")

        self.assertEqual(result.covered_intents, [])
        self.assertEqual(len(result.warnings), 1)
        self.assertIn("unknown intent", result.warnings[0])

    def test_run_returns_warning_for_invalid_llm_json(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["SummarizeTextToShorterVersion"],
            intent_tool_ids={"SummarizeTextToShorterVersion": ["Text Summarizer"]},
            llm_client=RawFakeLLMClient(
                '{"covered_intents":[{"intent":"SummarizeTextToShorterVersion" "coverage_type":"direct"}]}'
            ),
        )

        result = experiment.run("Summarize this text.")

        self.assertEqual(result.covered_intents, [])
        self.assertIn("LLM_COVERAGE_PARSE_FAILED", result.warnings[0])
        self.assertIn("JSONDecodeError", result.warnings[0])
        self.assertIn("raw_text", result.raw_llm_output)

    def test_run_reraises_fatal_llm_setup_errors(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["SummarizeTextToShorterVersion"],
            llm_client=RaisingLLMClient(FileNotFoundError("llm config file not found")),
        )

        with self.assertRaises(FileNotFoundError):
            experiment.run("Summarize this text.")

    def test_cli_fails_fast_for_missing_llm_config(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            missing_config = Path(tmpdir) / "missing.json"

            with self.assertRaises(SystemExit) as raised:
                main(
                    [
                        "--request",
                        "Summarize this text.",
                        "--llm-config",
                        str(missing_config),
                    ]
                )

        self.assertEqual(raised.exception.code, 2)

    def test_prompt_covers_non_english_content_level_operations(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["TranslateTextToEnglish"],
            coverage_prompt_rules=MULTIMEDIA_COVERAGE_PROMPT_RULES,
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt("I have a paragraph written in French and need images.")

        self.assertIn("written in French", prompt)
        self.assertIn("content-level operation", prompt)
        self.assertIn("explicitly says the input text", prompt)
        self.assertIn("written in another language", prompt)
        self.assertIn("result in a different", prompt)
        self.assertIn("understanding", prompt)
        self.assertIn("strongly_implied", prompt)

    def test_prompt_covers_image_visualization_phrases(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["GenerateImageByDescription", "SearchImagesByQuery"],
            coverage_prompt_rules=MULTIMEDIA_COVERAGE_PROMPT_RULES,
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt("I need some images that can visualize the sentiment.")

        self.assertIn("need some images", prompt)
        self.assertIn("visualize", prompt)
        self.assertIn("image generation", prompt)
        self.assertIn("image search", prompt)

    def test_prompt_covers_url_download_for_media_sources(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["DownloadVideoFromURL", "ExtractAudioFromVideo"],
            coverage_prompt_rules=MULTIMEDIA_COVERAGE_PROMPT_RULES,
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt("Extract speech from the video at https://example.com/video.mp4.")

        self.assertIn("URL phrases decompose by role", prompt)
        self.assertIn("direct resource", prompt)
        self.assertIn("embedded link in text", prompt)
        self.assertIn("link extraction target", prompt)
        self.assertIn("not execution-order inference", prompt)
        self.assertIn("source resource to process", prompt)
        self.assertIn("matching URL download intent", prompt)
        self.assertIn("embedded inside a larger text/message", prompt)
        self.assertIn("URL extraction plus the matching URL download intent", prompt)
        self.assertIn("strongly_implied", prompt)
        self.assertIn("Cover URL extraction alone only when", prompt)
        self.assertIn("URLs/links themselves", prompt)

    def test_prompt_uses_tool_metadata_for_url_download_disambiguation(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["FetchVideoAssetFromLink", "ExtractAudioFromVideo", "SearchImagebyImage"],
            intent_tool_ids={
                "FetchVideoAssetFromLink": ["Any Video URL Tool"],
                "ExtractAudioFromVideo": ["Video-to-Audio"],
                "SearchImagebyImage": ["Image Search (by Image)"],
            },
            intent_tool_descs={
                "FetchVideoAssetFromLink": [
                    {
                        "tool_id": "Any Video URL Tool",
                        "desc": "Fetches a video file from a web link.",
                        "input_types": ["url"],
                        "output_types": ["video"],
                    }
                ]
            },
            coverage_prompt_rules=MULTIMEDIA_COVERAGE_PROMPT_RULES,
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt(
            "I'd like to create an audio file from the speech in the video at "
            "https://www.example.com/video.mp4. Then generate a waveform image "
            "and search for similar images using the generated image."
        )

        self.assertIn("FetchVideoAssetFromLink", prompt)
        self.assertIn("Any Video URL Tool", prompt)
        self.assertIn('"input_types": [', prompt)
        self.assertIn('"url"', prompt)
        self.assertIn('"output_types": [', prompt)
        self.assertIn('"video"', prompt)
        self.assertIn("matching URL download intent", prompt)
        self.assertIn("url_download_candidates", prompt)
        self.assertIn("source resource to process", prompt)

    def test_run_does_not_apply_hardcoded_intent_augmentations(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=[
                "DownloadAudioFromURL",
                "TranslateTextToEnglish",
                "ApplyEffectToAudioByInstruction",
                "RewriteAndParaphraseTextUsingDifferentWords",
                "RewriteArticleOrSpinnerArticleUsingSynonymsOrSyntax",
            ],
            intent_tool_ids={
                "DownloadAudioFromURL": ["Audio Downloader"],
                "TranslateTextToEnglish": ["Text Translator"],
                "ApplyEffectToAudioByInstruction": ["Audio Effects"],
                "RewriteAndParaphraseTextUsingDifferentWords": ["Text Paraphraser"],
                "RewriteArticleOrSpinnerArticleUsingSynonymsOrSyntax": ["Article Spinner"],
            },
            llm_client=FakeLLMClient(
                {
                    "covered_intents": [
                        {
                            "intent": "ApplyEffectToAudioByInstruction",
                            "coverage_type": "synonym",
                            "confidence": 0.86,
                            "matched_request_phrase": "apply bajolaunch",
                            "matched_intent_term": "apply audio effect by instruction",
                            "reason": "The request asks to apply a chorus-like audio effect.",
                        },
                    ]
                }
            ),
        )

        result = experiment.run(
            "Use https://www.example.com/example_audio.wav. My instruction is in Portuguese: "
            "'Aplique bajolaunch com moderada intensidade no audio.' Also rewrite an article to avoid plagiarism."
        )

        self.assertEqual(
            [row.intent for row in result.covered_intents],
            ["ApplyEffectToAudioByInstruction"],
        )
        self.assertEqual(result.covered_intents[0].tool_ids, ["Audio Effects"])
        self.assertEqual(result.covered_intents[0].coverage_type, "synonym")
        self.assertNotIn(
            "DownloadAudioFromURL",
            [row.intent for row in result.covered_intents],
        )
        self.assertNotIn("TranslateTextToEnglish", [row.intent for row in result.covered_intents])
        self.assertNotIn(
            "RewriteArticleOrSpinnerArticleUsingSynonymsOrSyntax",
            [row.intent for row in result.covered_intents],
        )

    def test_prompt_requires_independent_coverage_for_coordinated_actions(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=[
                "RewriteAndParaphraseTextUsingDifferentWords",
                "RewriteAndSimplifyTextToBeUnderstandable",
                "SearchImagesByQuery",
            ],
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt("I would like it to be paraphrased, simplified, and then find an image.")

        self.assertIn("Decompose the request into separate action phrases", prompt)
        self.assertIn("Evaluate every allowed intent independently", prompt)
        self.assertIn("Do not collapse two matched intents into one", prompt)
        self.assertIn("return all of", prompt)

    def test_prompt_requires_verbatim_allowed_intent_names(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["AnswerQuestionFromNaturalImage"],
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt("Answer the question from the image.")

        self.assertIn("covered_intents[*].intent must be copied verbatim", prompt)
        self.assertIn("allowed_intents[*].intent", prompt)
        self.assertIn("Do not invent, paraphrase, shorten, translate, or", prompt)
        self.assertIn("delete any row whose intent", prompt)
        self.assertIn('"intent": "copy one exact allowed_intents[*].intent value"', prompt)

    def test_default_prompt_does_not_embed_multimedia_specific_rules(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["RewriteAndParaphraseTextUsingDifferentWords", "DownloadVideoFromURL"],
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt("Rewrite the article at https://example.com/a.txt.")

        self.assertNotIn("article-level objects", prompt)
        self.assertNotIn("matching URL download intent", prompt)
        self.assertNotIn("image generation", prompt)

    def test_prompt_uses_tool_descriptions_for_rewrite_disambiguation(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=[
                "RewriteAndParaphraseTextUsingDifferentWords",
                "RewriteArticleOrSpinnerArticleUsingSynonymsOrSyntax",
            ],
            intent_tool_ids={
                "RewriteAndParaphraseTextUsingDifferentWords": ["Text Paraphraser"],
                "RewriteArticleOrSpinnerArticleUsingSynonymsOrSyntax": ["Article Spinner"],
            },
            intent_tool_descs={
                "RewriteAndParaphraseTextUsingDifferentWords": [
                    {
                        "tool_id": "Text Paraphraser",
                        "desc": "Rewrites a given text using different words while maintaining its original meaning.",
                    }
                ],
                "RewriteArticleOrSpinnerArticleUsingSynonymsOrSyntax": [
                    {
                        "tool_id": "Article Spinner",
                        "desc": "Rewrites a given article using synonyms and syntax changes to create a new, unique version.",
                    }
                ],
            },
            coverage_prompt_rules=MULTIMEDIA_COVERAGE_PROMPT_RULES,
            llm_client=FakeLLMClient({"covered_intents": []}),
        )

        prompt = experiment._build_prompt("Rewrite an article 'example.txt' into a unique version.")

        self.assertIn("candidate_tools", prompt)
        self.assertIn("Article Spinner", prompt)
        self.assertIn("synonyms and syntax changes", prompt)
        self.assertIn("Text Paraphraser", prompt)
        self.assertIn("different words", prompt)
        self.assertIn("Use candidate tool descriptions only as evidence", prompt)
        self.assertIn("article-level objects", prompt)
        self.assertIn("text-level objects", prompt)

    def test_loads_unique_intents_from_tool_desc(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tool_desc.json"
            path.write_text(json.dumps(_sample_tool_desc()), encoding="utf-8")

            intents = load_intents(path)
            intent_tool_ids = load_intent_tool_ids(path)
            intent_tools = load_intent_tools(path)

        self.assertEqual(
            intents,
            [
                "SummarizeTextToShorterVersion",
                "AnalysisSentimentOfText",
            ],
        )
        self.assertEqual(
            intent_tool_ids["SummarizeTextToShorterVersion"],
            ["Text Summarizer", "Another Summarizer"],
        )
        self.assertEqual(
            intent_tools["SummarizeTextToShorterVersion"][0]["desc"],
            "Summarizes text.",
        )

    def test_batch_rows_include_qwen_result_and_tools(self) -> None:
        experiment = IntentLLMCoverageExperiment(
            intents=["SummarizeTextToShorterVersion"],
            intent_tool_ids={"SummarizeTextToShorterVersion": ["Text Summarizer"]},
            llm_client=FakeLLMClient(
                {
                    "covered_intents": [
                        {
                            "intent": "SummarizeTextToShorterVersion",
                            "coverage_type": "direct",
                            "confidence": 0.9,
                            "matched_request_phrase": "summary",
                            "matched_intent_term": "summarize",
                            "reason": "The request asks for a summary.",
                        }
                    ]
                }
            ),
        )

        result = experiment.run("Give me a summary.")
        rows = coverage_result_to_batch_rows(
            case_id="case-1",
            user_request="Give me a summary.",
            gold_result={"tool_nodes": [{"task": "Text Summarizer"}]},
            qwen_result={"task_nodes": [{"task": "Text Summarizer"}]},
            result=result,
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(list(rows[0]), BATCH_COLUMNS)
        self.assertEqual(rows[0]["id"], "case-1")
        self.assertEqual(rows[0]["user_request"], "Give me a summary.")
        self.assertEqual(rows[0]["intent"], "SummarizeTextToShorterVersion")
        self.assertEqual(rows[0]["intent tool"], "Text Summarizer")
        self.assertEqual(rows[0]["model tool"], "Text Summarizer")
        self.assertEqual(rows[0]["coverage_warnings"], "")
        self.assertNotIn("tools", rows[0])
        self.assertNotIn("gold tool", rows[0])

    def test_loads_jsonl_qwen_results_and_writes_csv(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            case_file = tmp / "cases.jsonl"
            qwen_file = tmp / "qwen.json"
            gold_file = tmp / "data.json"
            output_file = tmp / "coverage.csv"
            case_file.write_text(
                json.dumps({"id": "1", "user_request": "Summarize this."}) + "\n",
                encoding="utf-8",
            )
            qwen_file.write_text(
                json.dumps({"id": "1", "result": {"task_nodes": [{"task": "Text Summarizer"}]}}) + "\n",
                encoding="utf-8",
            )
            gold_file.write_text(
                json.dumps(
                    {
                        "id": "1",
                        "tool_steps": json.dumps(["Step 1: Summarize text."]),
                        "tool_nodes": json.dumps([{"task": "Text Summarizer"}]),
                        "tool_links": "[]",
                    }
                )
                + "\n",
                encoding="utf-8",
            )

            cases = load_json_records(case_file)
            gold_results = load_gold_results(gold_file)
            qwen_results = load_qwen_results(qwen_file)
            rows = [
                {
                    "id": "1",
                    "user_request": cases[0]["user_request"],
                    "intent": "SummarizeTextToShorterVersion",
                    "intent tool": "Text Summarizer",
                    "model tool": "Text Summarizer",
                    "coverage_warnings": "",
                }
            ]
            write_batch_table(rows, output_file, "csv")

            csv_text = output_file.read_text(encoding="utf-8-sig")
            csv_rows = list(csv.DictReader(io.StringIO(csv_text)))

        self.assertEqual(cases[0]["id"], "1")
        self.assertEqual(gold_results["1"]["tool_nodes"][0]["task"], "Text Summarizer")
        self.assertIn("task_nodes", json.dumps(qwen_results["1"]))
        self.assertIn("id,user_request,intent,intent tool,model tool,coverage_warnings", csv_text)
        self.assertNotIn("gold tool", csv_text)
        self.assertNotIn("gold_result", csv_text)
        self.assertIn("SummarizeTextToShorterVersion", csv_text)
        self.assertEqual(csv_rows[0]["intent tool"], "Text Summarizer")
        self.assertNotIn("request匹配片段", csv_text)

    def test_comparison_row_extracts_gold_qwen_and_gpt_tools(self) -> None:
        result = IntentLLMCoverageExperiment(
            intents=["SummarizeTextToShorterVersion"],
            intent_tool_ids={"SummarizeTextToShorterVersion": ["Text Summarizer"]},
            llm_client=FakeLLMClient(
                {
                    "covered_intents": [
                        {
                            "intent": "SummarizeTextToShorterVersion",
                            "coverage_type": "direct",
                            "confidence": 0.9,
                            "matched_request_phrase": "summary",
                            "matched_intent_term": "summarize",
                            "reason": "The request asks for a summary.",
                        }
                    ]
                }
            ),
        ).run("Give me a summary.")

        row = coverage_result_to_comparison_row(
            case={"type": "chain", "badcase_error_summary": "node_mismatch"},
            case_id="case-1",
            user_request="Give me a summary.",
            gold_result={"tool_nodes": [{"task": "Text Summarizer"}]},
            qwen_result={"task_nodes": [{"task": "Text Search"}, {"task": "Text Summarizer"}]},
            result=result,
        )

        self.assertEqual(list(row), COMPARISON_TABLE_COLUMNS)
        self.assertEqual(row["model tool"], "Text Search -> Text Summarizer")
        self.assertEqual(row["intent tool"], "Text Summarizer")
        self.assertEqual(row["intent"], "SummarizeTextToShorterVersion")
        self.assertEqual(row["coverage_warnings"], "")

    def test_run_comparison_table_batch_and_write_markdown(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            case_file = tmp / "cases.jsonl"
            output_file = tmp / "coverage_table.md"
            case_file.write_text(
                json.dumps(
                    {
                        "id": "1",
                        "type": "dag",
                        "user_request": "First request: summary.",
                        "gold_result": {"tool_nodes": [{"task": "Text Summarizer"}]},
                        "qwen_result": {"task_nodes": [{"task": "Text Summarizer"}]},
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            experiment = IntentLLMCoverageExperiment(
                intents=["SummarizeTextToShorterVersion"],
                intent_tool_ids={"SummarizeTextToShorterVersion": ["Text Summarizer"]},
                llm_client=FakeLLMClient(
                    {
                        "covered_intents": [
                            {
                                "intent": "SummarizeTextToShorterVersion",
                                "coverage_type": "direct",
                                "confidence": 0.9,
                                "matched_request_phrase": "summary",
                                "matched_intent_term": "summarize",
                                "reason": "The request asks for a summary.",
                            }
                        ]
                    }
                ),
            )

            rows = run_comparison_table_batch(experiment=experiment, input_file=case_file)
            write_comparison_table(rows, output_file, "table-md")
            markdown = output_file.read_text(encoding="utf-8")

        self.assertEqual(rows[0]["user_request"], "First request: summary.")
        self.assertIn(
            "| " + " | ".join(COMPARISON_TABLE_COLUMNS) + " |",
            markdown,
        )
        self.assertIn("Text Summarizer", markdown)

    def test_comparison_table_records_invalid_llm_json_warning(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            case_file = Path(tmpdir) / "cases.jsonl"
            case_file.write_text(
                json.dumps(
                    {
                        "id": "1",
                        "type": "dag",
                        "user_request": "First request: summary.",
                        "gold_result": {"tool_nodes": [{"task": "Text Summarizer"}]},
                        "qwen_result": {"task_nodes": [{"task": "Text Search"}]},
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            experiment = IntentLLMCoverageExperiment(
                intents=["SummarizeTextToShorterVersion"],
                intent_tool_ids={"SummarizeTextToShorterVersion": ["Text Summarizer"]},
                llm_client=RawFakeLLMClient(
                    '{"covered_intents":[{"intent":"SummarizeTextToShorterVersion" "coverage_type":"direct"}]}'
                ),
            )

            rows = run_comparison_table_batch(experiment=experiment, input_file=case_file, workers=2)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["model tool"], "Text Search")
        self.assertEqual(rows[0]["intent tool"], "")
        self.assertEqual(rows[0]["intent"], "")
        self.assertIn("LLM_COVERAGE_PARSE_FAILED", rows[0]["coverage_warnings"])
        self.assertIn("JSONDecodeError", rows[0]["coverage_warnings"])

    def test_comparison_table_resume_skips_success_and_reruns_failed(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            case_file = tmp / "cases.jsonl"
            checkpoint_file = tmp / "coverage.xlsx.checkpoint.jsonl"
            case_file.write_text(
                "\n".join(
                    [
                        json.dumps(
                            {
                                "id": "1",
                                "type": "dag",
                                "user_request": "First request: summary.",
                                "gold_result": {"tool_nodes": [{"task": "Text Summarizer"}]},
                                "qwen_result": {"task_nodes": [{"task": "Text Summarizer"}]},
                            }
                        ),
                        json.dumps(
                            {
                                "id": "2",
                                "type": "dag",
                                "user_request": "Second request: sentiment.",
                                "gold_result": {"tool_nodes": [{"task": "Text Sentiment Analysis"}]},
                                "qwen_result": {"task_nodes": [{"task": "Text Search"}]},
                            }
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            completed_row = {
                "id": "1",
                "user_request": "First request: summary.",
                "intent": "SummarizeTextToShorterVersion",
                "intent tool": "Text Summarizer",
                "model tool": "Text Summarizer",
                "coverage_warnings": "",
            }
            failed_row = {
                "id": "2",
                "user_request": "Second request: sentiment.",
                "intent": "",
                "intent tool": "",
                "model tool": "Text Search",
                "coverage_warnings": "LLM_COVERAGE_PARSE_FAILED: JSONDecodeError: bad json",
            }
            checkpoint_file.write_text(
                json.dumps(
                    {
                        "mode": "comparison",
                        "key": "1",
                        "position": 0,
                        "id": "1",
                        "status": "ok",
                        "row": completed_row,
                    },
                    ensure_ascii=False,
                )
                + "\n"
                + json.dumps(
                    {
                        "mode": "comparison",
                        "key": "2",
                        "position": 1,
                        "id": "2",
                        "status": "failed",
                        "row": failed_row,
                    },
                    ensure_ascii=False,
                )
                + "\n",
                encoding="utf-8",
            )
            client = CountingPromptAwareFakeLLMClient()
            experiment = IntentLLMCoverageExperiment(
                intents=["SummarizeTextToShorterVersion", "AnalysisSentimentOfText"],
                intent_tool_ids={
                    "SummarizeTextToShorterVersion": ["Text Summarizer"],
                    "AnalysisSentimentOfText": ["Text Sentiment Analysis"],
                },
                llm_client=client,
            )

            rows = run_comparison_table_batch(
                experiment=experiment,
                input_file=case_file,
                workers=2,
                checkpoint_file=checkpoint_file,
                resume=True,
            )
            checkpoint_records = [
                json.loads(line)
                for line in checkpoint_file.read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]

        self.assertEqual(len(client.prompts), 1)
        self.assertIn("Second request", client.prompts[0])
        self.assertEqual([row["id"] for row in rows], ["1", "2"])
        self.assertEqual(rows[0]["intent"], "SummarizeTextToShorterVersion")
        self.assertEqual(rows[1]["intent"], "AnalysisSentimentOfText")
        self.assertEqual(checkpoint_records[-1]["id"], "2")
        self.assertEqual(checkpoint_records[-1]["status"], "ok")

    def test_user_requests_file_overrides_instruction_and_existing_coverage_skips(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            case_file = tmp / "data.json"
            user_requests_file = tmp / "user_requests.json"
            existing_coverage_file = tmp / "existing_coverage.csv"
            case_file.write_text(
                "\n".join(
                    [
                        json.dumps({"id": "1", "instruction": "Instruction summary should not be used."}),
                        json.dumps({"id": "2", "instruction": "Instruction sentiment should not be used."}),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            user_requests_file.write_text(
                "\n".join(
                    [
                        json.dumps({"id": "1", "user_request": "First request: summary canonical."}),
                        json.dumps({"id": "2", "user_request": "Second request: sentiment canonical."}),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            write_comparison_table(
                [
                    {
                        "id": "1",
                        "user_request": "stale request text",
                        "intent": "SummarizeTextToShorterVersion",
                        "intent tool": "Text Summarizer",
                        "model tool": "",
                        "coverage_warnings": "",
                    },
                    {
                        "id": "2",
                        "user_request": "stale failed request text",
                        "intent": "",
                        "intent tool": "",
                        "model tool": "",
                        "coverage_warnings": "LLM_COVERAGE_PARSE_FAILED: JSONDecodeError: bad json",
                    },
                ],
                existing_coverage_file,
                "csv",
            )
            client = CountingPromptAwareFakeLLMClient()
            experiment = IntentLLMCoverageExperiment(
                intents=["SummarizeTextToShorterVersion", "AnalysisSentimentOfText"],
                intent_tool_ids={
                    "SummarizeTextToShorterVersion": ["Text Summarizer"],
                    "AnalysisSentimentOfText": ["Text Sentiment Analysis"],
                },
                llm_client=client,
            )

            rows = run_comparison_table_batch(
                experiment=experiment,
                input_file=case_file,
                user_request_by_id=load_user_requests(user_requests_file),
                existing_coverage_by_id=load_existing_coverage_tables([str(existing_coverage_file)]),
            )

        self.assertEqual(len(client.prompts), 1)
        self.assertIn("Second request: sentiment canonical.", client.prompts[0])
        self.assertNotIn("Instruction sentiment should not be used.", client.prompts[0])
        self.assertEqual(rows[0]["user_request"], "First request: summary canonical.")
        self.assertEqual(rows[0]["intent"], "SummarizeTextToShorterVersion")
        self.assertEqual(rows[1]["user_request"], "Second request: sentiment canonical.")
        self.assertEqual(rows[1]["intent"], "AnalysisSentimentOfText")

    def test_write_comparison_table_xlsx(self) -> None:
        rows = [
            {
                "id": "case-1",
                "user_request": "Give me a summary.",
                "intent": "SummarizeTextToShorterVersion",
                "intent tool": "Text Summarizer",
                "model tool": "Text Search",
                "coverage_warnings": "",
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = Path(tmpdir) / "coverage_table.xlsx"
            try:
                import openpyxl
            except ImportError:
                with self.assertRaisesRegex(RuntimeError, "pip install openpyxl"):
                    write_comparison_table(rows, output_file, "table-xlsx")
                return

            write_comparison_table(rows, output_file, "table-xlsx")
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]

        self.assertEqual(worksheet.cell(row=1, column=1).value, "id")
        self.assertEqual(worksheet.cell(row=2, column=1).value, "case-1")
        self.assertEqual(
            worksheet.cell(row=2, column=headers.index("intent") + 1).value,
            "SummarizeTextToShorterVersion",
        )

    def test_run_batch_can_use_workers_and_preserves_case_order(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            case_file = Path(tmpdir) / "cases.jsonl"
            case_file.write_text(
                "\n".join(
                    [
                        json.dumps({"id": "1", "user_request": "First request: summary."}),
                        json.dumps({"id": "2", "user_request": "Second request: sentiment."}),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            experiment = IntentLLMCoverageExperiment(
                intents=["SummarizeTextToShorterVersion", "AnalysisSentimentOfText"],
                intent_tool_ids={
                    "SummarizeTextToShorterVersion": ["Text Summarizer"],
                    "AnalysisSentimentOfText": ["Text Sentiment Analysis"],
                },
                llm_client=PromptAwareFakeLLMClient(),
            )

            rows = run_batch(experiment=experiment, input_file=case_file, workers=2)

        self.assertEqual([row["id"] for row in rows], ["1", "2"])
        self.assertEqual(rows[0]["intent"], "SummarizeTextToShorterVersion")
        self.assertEqual(rows[1]["intent tool"], "Text Sentiment Analysis")

    def test_loads_bom_encoded_json_and_jsonl(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            tool_desc_file = tmp / "tool_desc.json"
            records_file = tmp / "records.jsonl"
            tool_desc_file.write_text(json.dumps(_sample_tool_desc()), encoding="utf-8-sig")
            records_file.write_text(
                json.dumps({"id": "1", "user_request": "Summarize this."}) + "\n",
                encoding="utf-8-sig",
            )

            intents = load_intents(tool_desc_file)
            records = load_json_records(records_file)

        self.assertEqual(intents[0], "SummarizeTextToShorterVersion")
        self.assertEqual(records[0]["id"], "1")


def _sample_tool_desc():
    return {
        "nodes": [
            {
                "id": "Text Summarizer",
                "intent": "SummarizeTextToShorterVersion",
                "desc": "Summarizes text.",
            },
            {
                "id": "Another Summarizer",
                "intent": "SummarizeTextToShorterVersion",
                "desc": "Also summarizes text.",
            },
            {
                "id": "Text Sentiment Analysis",
                "intent": "AnalysisSentimentOfText",
                "desc": "Analyzes text sentiment.",
            },
        ]
    }


if __name__ == "__main__":
    unittest.main()
